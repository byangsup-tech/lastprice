from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.models import Product

HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(bold=True, color="FFFFFF")
CENTER = Alignment(horizontal="center", vertical="center")

HEADERS = ["회사", "상품코드", "상품명", "성별", "연령", "보험기간", "납입기간",
           "특약명", "비고", "최저가입금액", "최고가입금액", "설계금액",
           "납기", "만기", "월보험료(원)", "수집일시"]
AMOUNT_COLS = (10, 11, 12, 15)   # 최저·최고·설계금액·월보험료 — 천단위 구분
COL_WIDTHS = [12, 10, 40, 6, 6, 10, 10, 44, 10, 13, 13, 13, 10, 10, 13, 20]


def write_long_workbook(company: str, products: list[Product], output_dir: Path) -> Path:
    """조건×특약 long-format 워크북 — 1행=1관측(피벗용 평면표).

    products 는 (상품 × 조건프로파일) 단위 Product 리스트. Product 하나가
    한 조건에서의 한 상품 산출결과이고, 그 안의 특약마다 1행을 쓴다. 동일 특약을
    성별·연령별로 비교하려면 이 시트를 피벗 테이블 소스로 쓰면 된다.
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "보험료"

    for c, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER

    r = 1
    for product in products:
        captured = product.captured_at.strftime("%Y-%m-%d %H:%M:%S")
        base = [company, product.code or "", product.name, product.cond_sex,
                _as_int(product.cond_age), product.cond_insurance_period,
                product.cond_payment_period]
        # 특약 0건(수집 실패)이면 조건 식별용 1행이라도 남긴다
        for rider in (product.riders or [None]):
            r += 1
            if rider is None:
                tail = ["(수집 실패)", product.error or "",
                        None, None, None, "", "", None, captured]
            else:
                tail = [rider.name, rider.note, rider.min_amount, rider.max_amount,
                        rider.selected_amount, rider.pay_period, rider.maturity,
                        rider.premium, captured]
            for c, v in enumerate(base + tail, start=1):
                cell = ws.cell(row=r, column=c, value=v)
                if c in AMOUNT_COLS and isinstance(v, int):
                    cell.number_format = "#,##0"

    if r == 1:
        ws.cell(row=2, column=1, value=f"{company}: 수집 결과 없음")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{max(r, 2)}"
    for i, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    out_path = output_dir / f"{_safe_filename(company)}.xlsx"
    wb.save(out_path)
    return out_path


def _as_int(val):
    """'40' → 40. 변환 불가하면 원본(빈 문자열 포함) 그대로."""
    try:
        return int(str(val).strip())
    except (ValueError, TypeError):
        return val or ""


def _safe_filename(name: str) -> str:
    bad = '<>:"/\\|?*'
    return "".join("_" if c in bad else c for c in name).strip() or "company"
