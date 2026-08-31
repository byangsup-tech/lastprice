# -*- coding: utf-8 -*-
"""raw/ → CSV·XLSX.

emit 은 raw/ 와 사이드카의 순수 함수다. 네트워크도 키도 필요 없고, 몇 번을 돌려도
같은 결과가 나온다. 그래서 파서를 고친 뒤 쿼터를 다시 쓰지 않고 재조립할 수 있고,
중단된 실행은 '행이 적은 CSV + 전부 설명된 미확보목록' 이 된다 (중복도 축소도 아니다).
"""
from __future__ import annotations

import csv
import json
import os
import re
from datetime import datetime

import config
import corpcode
import docparse
import grid

PROV_COLS = ["source_endpoint", "source_params", "rcept_no", "rcept_no_source",
             "fetched_at", "status", "call_id", "raw_path", "raw_sha256", "data_age_days"]

# 금액 컬럼 판별 (union-of-keys 라 하드코딩 목록은 쓰지 않는다)
TERM_RE = re.compile(r"^(thstrm|thstrm_add|frmtrm_q|frmtrm_add|frmtrm|bfefrmtrm)_amount$")
PERIOD_TYPE = {
    "thstrm": "current", "thstrm_add": "current_cumulative",
    "frmtrm": "prior", "frmtrm_q": "prior_quarter", "frmtrm_add": "prior_cumulative",
    "bfefrmtrm": "before_prior",
}


# ── 원본 스캔 ─────────────────────────────────────────────────────────────
def scan_raw(out_dir, endpoint=None):
    """raw/<endpoint>/*.meta.json 을 훑는다. _transient/_quarantine 는 제외."""
    root = os.path.join(out_dir, "raw")
    if not os.path.isdir(root):
        return []
    metas = []
    for ep in sorted(os.listdir(root)):
        if ep.startswith("_"):
            continue
        if endpoint and ep != endpoint:
            continue
        d = os.path.join(root, ep)
        if not os.path.isdir(d):
            continue
        for fn in sorted(os.listdir(d)):
            if not fn.endswith(".meta.json"):
                continue
            try:
                with open(os.path.join(d, fn), encoding="utf-8") as f:
                    m = json.load(f)
            except Exception:
                continue
            m["_meta_path"] = os.path.join(d, fn)
            m["_body_path"] = os.path.join(d, fn[:-len(".meta.json")])
            m["_endpoint"] = m.get("endpoint", ep)
            metas.append(m)
    return metas


def load_body(meta):
    with open(meta["_body_path"], "rb") as f:
        return f.read()


def load_json(meta):
    try:
        return json.loads(load_body(meta).decode("utf-8-sig"))
    except Exception:
        return None


def rel(out_dir, p):
    return os.path.relpath(p, out_dir).replace(os.sep, "/")


def prov(out_dir, meta, rcept_no="", rcept_no_source="none"):
    age = ""
    try:
        age = (datetime.now().astimezone()
               - datetime.fromisoformat(meta["fetched_at"])).days
    except Exception:
        pass
    return {
        "source_endpoint": meta.get("endpoint", ""),
        "source_params": json.dumps(meta.get("params", {}), ensure_ascii=False, sort_keys=True),
        "rcept_no": rcept_no,
        "rcept_no_source": rcept_no_source,
        "fetched_at": meta.get("fetched_at", ""),
        "status": meta.get("api_status", ""),
        "call_id": meta.get("call_id", ""),
        "raw_path": rel(out_dir, meta["_body_path"]),
        "raw_sha256": meta.get("sha256", ""),
        "data_age_days": age,
    }


# ── 값 파싱 ───────────────────────────────────────────────────────────────
def parse_amount(row, key):
    """(원문, 숫자, parse_status). 빈칸의 의미를 뭉개지 않고, 절대 0 으로 채우지 않는다."""
    if key not in row:
        return "", "", "key_absent"          # DART 는 null 키를 아예 뺀다
    raw = row.get(key)
    if raw is None:
        return "", "", "key_absent"
    s = str(raw).strip()
    if s == "":
        return s, "", "empty_string"
    if s in ("-", "－", "—"):
        return s, "", "dash"
    t = s.replace(",", "").replace(" ", "")
    neg = False
    if t.startswith("△") or t.startswith("▲"):
        neg, t = True, t[1:]
    if t.startswith("(") and t.endswith(")"):
        neg, t = True, t[1:-1]
    if t.startswith("-"):
        neg, t = True, t[1:]
    try:
        v = float(t)
    except ValueError:
        return s, "", "unparseable"
    v = -v if neg else v
    return s, ("%d" % v if v == int(v) else repr(v)), "ok"


def accounting_std(bsns_year):
    try:
        y = int(bsns_year)
    except Exception:
        return "", ""
    rule = "보고서 사업연도 >= %d 이면 IFRS17, 아니면 IFRS4" % config.IFRS17_FIRST_YEAR
    return ("IFRS17" if y >= config.IFRS17_FIRST_YEAR else "IFRS4"), rule


_EVENTS_BY_LABEL = {}
for _lb, _d, _ty, _sv, _n in config.DISCONTINUITY_EVENTS:
    _EVENTS_BY_LABEL.setdefault(_lb, []).append((_d, _ty, _sv, _n))


def comparability(label, bsns_year):
    """행이 덮는 기간(당기·전기·전전기)에 단절 이벤트가 걸리면 표시만 한다."""
    try:
        y = int(bsns_year)
    except Exception:
        return "", ""
    hits = [(d, t, n) for d, t, s, n in _EVENTS_BY_LABEL.get(label, [])
            if y - 2 <= int(d[:4]) <= y]
    if int(bsns_year or 0) >= config.IFRS17_FIRST_YEAR > int(bsns_year or 0) - 2:
        hits.append((str(config.IFRS17_FIRST_YEAR) + "0101", "ifrs17", "IFRS4→IFRS17 전환"))
    if not hits:
        return "N", ""
    return "Y", " | ".join("%s %s(%s)" % (d, t, n) for d, t, n in hits)


# ── CSV 쓰기 ──────────────────────────────────────────────────────────────
def write_csv(path, rows, lead_cols=()):
    """컬럼은 관측된 키의 합집합. 고정 목록을 쓰지 않으므로 키 손실이 구조적으로 없다."""
    if not rows:
        return None, []
    seen, cols = set(), []
    for c in list(lead_cols):
        if c not in seen:
            seen.add(c); cols.append(c)
    extra = set()
    for r in rows:
        extra |= set(r.keys())
    for c in sorted(extra - seen - set(PROV_COLS)):
        cols.append(c); seen.add(c)
    for c in PROV_COLS:
        if c not in seen:
            cols.append(c); seen.add(c)
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow(r)
    return path, cols


# ── 엔드포인트별 산출 ─────────────────────────────────────────────────────
def corp_lookup(out_dir):
    entries = corpcode.load_corp_codes(out_dir)
    return {e["corp_code"]: e for e in entries}, entries


def emit_company(out_dir, metas, by_code):
    rows = []
    for m in metas:
        d = load_json(m) or {}
        cc = (m.get("params") or {}).get("corp_code", "")
        e = by_code.get(cc, {})
        row = {k: v for k, v in d.items() if k not in ("status", "message", "list")}
        row.update(corp_label=e.get("label", ""), entity_type=e.get("entity_type", ""),
                   group=e.get("group", ""), status_note=e.get("status_note", ""),
                   value_source="api", api_message=d.get("message", ""))
        row.update(prov(out_dir, m))
        rows.append(row)
    rows.sort(key=lambda r: (r.get("corp_label") or "", r.get("corp_code") or ""))
    return write_csv(os.path.join(out_dir, "01_기업개황.csv"), rows,
                     ["corp_label", "entity_type", "group", "corp_code", "corp_name",
                      "stock_code", "corp_cls", "est_dt", "jurir_no", "bizr_no",
                      "acc_mt", "adres", "induty_code", "ceo_nm"])


def emit_generic(out_dir, endpoint, metas, by_code, csv_name):
    rows = []
    for m in metas:
        p = m.get("params") or {}
        cc = p.get("corp_code", "")
        e = by_code.get(cc, {})
        std, rule = accounting_std(p.get("bsns_year"))
        for r in (load_json(m) or {}).get("list") or []:
            row = dict(r)
            rcept = str(r.get("rcept_no") or "")
            row.update(corp_label=e.get("label", ""), entity_type=e.get("entity_type", ""),
                       group=e.get("group", ""),
                       bsns_year=p.get("bsns_year", ""), reprt_code=p.get("reprt_code", ""),
                       reprt_nm=config.REPRT_NAMES.get(p.get("reprt_code", ""), ""),
                       accounting_std_inferred=std, inference_rule=rule,
                       value_source="api")
            row.update(prov(out_dir, m, rcept, "api_row" if rcept else "none"))
            rows.append(row)
    rows.sort(key=lambda r: (r.get("corp_label") or "", r.get("bsns_year") or "",
                             r.get("reprt_code") or ""))
    return write_csv(os.path.join(out_dir, csv_name + ".csv"), rows,
                     ["corp_label", "entity_type", "group", "corp_code", "corp_name",
                      "bsns_year", "reprt_code", "reprt_nm"])


def emit_list(out_dir, metas, by_code):
    rows, pagination = [], {}
    for m in metas:
        p = m.get("params") or {}
        cc = p.get("corp_code", "")
        e = by_code.get(cc, {})
        d = load_json(m) or {}
        key = (cc, p.get("bgn_de"), p.get("end_de"))
        pagination.setdefault(key, set()).add(str(d.get("total_count", "")))
        for r in d.get("list") or []:
            row = dict(r)
            row.update(corp_label=e.get("label", ""), entity_type=e.get("entity_type", ""),
                       group=e.get("group", ""), page_no=p.get("page_no", ""),
                       total_count=d.get("total_count", ""), total_page=d.get("total_page", ""),
                       report_nm_normalized=docparse.normalize_for_match(r.get("report_nm", "")),
                       value_source="api")
            row.update(prov(out_dir, m, str(r.get("rcept_no") or ""), "api_row"))
            rows.append(row)
    rows.sort(key=lambda r: (r.get("corp_label") or "", r.get("rcept_dt") or ""))
    path, _ = write_csv(os.path.join(out_dir, "02_공시목록.csv"), rows,
                        ["corp_label", "entity_type", "group", "corp_code", "corp_name",
                         "rcept_dt", "report_nm", "report_nm_normalized", "rcept_no",
                         "flr_nm", "rm", "corp_cls"])
    inconsistent = {k: v for k, v in pagination.items() if len(v) > 1}
    return path, rows, inconsistent


def emit_financials(out_dir, metas, by_code):
    wide, long_rows, accounts = [], [], {}
    drift = set()
    known = {"rcept_no", "reprt_code", "bsns_year", "corp_code", "sj_div", "sj_nm",
             "account_id", "account_nm", "account_detail", "ord", "currency",
             "thstrm_nm", "thstrm_amount", "thstrm_add_amount", "frmtrm_nm",
             "frmtrm_amount", "frmtrm_q_nm", "frmtrm_q_amount", "frmtrm_add_amount",
             "bfefrmtrm_nm", "bfefrmtrm_amount", "fs_div", "fs_nm"}
    for m in metas:
        p = m.get("params") or {}
        cc, fs_div = p.get("corp_code", ""), p.get("fs_div", "")
        by = p.get("bsns_year", "")
        e = by_code.get(cc, {})
        label = e.get("label", "")
        std, rule = accounting_std(by)
        brk, brk_note = comparability(label, by)
        file_rows = (load_json(m) or {}).get("list") or []
        # 파일 안에서 한 번이라도 관측된 term 은 그 파일의 '기대 형태'다. 어떤 행에 그 키가
        # 없으면 행을 빼는 게 아니라 key_absent 로 남긴다 — 빠뜨리면 '미보고'와 '미수집'을
        # 구분할 수 없게 되고, 그건 삭제나 마찬가지다.
        terms_in_file = []
        for r in file_rows:
            for k in r:
                mt = TERM_RE.match(k)
                if mt and mt.group(1) not in terms_in_file:
                    terms_in_file.append(mt.group(1))
        terms_in_file.sort(key=lambda t: list(PERIOD_TYPE).index(t) if t in PERIOD_TYPE else 99)
        for r in file_rows:
            drift |= (set(r.keys()) - known)
            rcept = str(r.get("rcept_no") or "")
            base = dict(
                corp_label=label, entity_type=e.get("entity_type", ""), group=e.get("group", ""),
                corp_code=cc, bsns_year=by, reprt_code=p.get("reprt_code", ""),
                reprt_nm=config.REPRT_NAMES.get(p.get("reprt_code", ""), ""),
                fs_div=fs_div, fs_div_requested=fs_div,
                acc_mt=e.get("acc_mt", ""),
                accounting_std_inferred=std, inference_rule=rule,
                comparability_break=brk, comparability_note=brk_note,
                currency=r.get("currency", ""), value_source="api")
            row = dict(r); row.update(base)
            row.update(prov(out_dir, m, rcept, "api_row" if rcept else "none"))
            wide.append(row)

            key = (label, fs_div, r.get("sj_div", ""), r.get("account_id", ""),
                   r.get("account_nm", ""), r.get("account_detail", ""), r.get("ord", ""))
            acc = accounts.setdefault(key, dict(
                corp_label=label, entity_type=e.get("entity_type", ""), fs_div=fs_div,
                sj_div=r.get("sj_div", ""), sj_nm=r.get("sj_nm", ""),
                account_id=r.get("account_id", ""), account_nm=r.get("account_nm", ""),
                account_detail=r.get("account_detail", ""), ord=r.get("ord", ""),
                account_id_is_standard="N" if (not r.get("account_id") or "미사용" in str(r.get("account_id"))) else "Y",
                years=set(), stds=set(), n_rows=0, value_source="api"))
            acc["years"].add(str(by)); acc["stds"].add(std); acc["n_rows"] += 1

            # 고정 3-term 가정을 쓰지 않는다. 파일에서 관측된 term 을 전부 편다.
            for term in terms_in_file:
                k = "%s_amount" % term
                raw, num, st = parse_amount(r, k)
                label_key = "%s_nm" % term
                term_label = r.get(label_key) or r.get("%s_nm" % term.split("_")[0]) or ""
                lr = dict(base)
                lr.update(sj_div=r.get("sj_div", ""), sj_nm=r.get("sj_nm", ""),
                          account_id=r.get("account_id", ""), account_nm=r.get("account_nm", ""),
                          account_detail=r.get("account_detail", ""), ord=r.get("ord", ""),
                          term_code=term, term_key=k, term_label=term_label,
                          period_type=PERIOD_TYPE.get(term, "unknown"),
                          is_comparative="N" if term.startswith("thstrm") else "Y",
                          amount_raw=raw, amount=num, parse_status=st)
                lr.update(prov(out_dir, m, rcept, "api_row" if rcept else "none"))
                long_rows.append(lr)

    lead_w = ["corp_label", "entity_type", "group", "corp_code", "bsns_year", "reprt_code",
              "reprt_nm", "fs_div", "accounting_std_inferred", "comparability_break",
              "sj_div", "sj_nm", "account_id", "account_nm", "account_detail", "ord", "currency"]
    p1, _ = write_csv(os.path.join(out_dir, "03_재무제표.csv"), wide, lead_w)
    lead_l = lead_w[:11] + ["term_code", "term_label", "period_type", "is_comparative",
                            "amount_raw", "amount", "parse_status"] + lead_w[11:]
    p2, _ = write_csv(os.path.join(out_dir, "03b_재무제표_long.csv"), long_rows, lead_l)

    arows = []
    for a in accounts.values():
        a = dict(a)
        a["years_seen"] = ",".join(sorted(a.pop("years")))
        a["accounting_std_seen"] = ",".join(sorted(x for x in a.pop("stds") if x))
        arows.append(a)
    arows.sort(key=lambda r: (r["corp_label"], r["fs_div"], r["sj_div"], str(r["ord"])))
    p3, _ = write_csv(os.path.join(out_dir, "00_계정과목목록.csv"), arows,
                      ["corp_label", "entity_type", "fs_div", "sj_div", "sj_nm",
                       "account_id", "account_id_is_standard", "account_nm",
                       "account_detail", "ord", "accounting_std_seen", "years_seen", "n_rows"])
    return [p1, p2, p3], sorted(drift)


def emit_ownership(out_dir, metas, by_code, entries):
    """★계열사 지분율 — 원표만 두지 않고 소유 그래프로 낸다."""
    name_to_label = {}
    for e in entries:
        name_to_label[corpcode.norm(e["label"])] = e
        for n in (e.get("corp_name_dart"), ):
            if n:
                name_to_label.setdefault(corpcode.norm(n), e)
    rows = []
    for m in metas:
        p = m.get("params") or {}
        cc = p.get("corp_code", "")
        parent = by_code.get(cc, {})
        for r in (load_json(m) or {}).get("list") or []:
            child_name = str(r.get("inv_prm") or "")
            hit = name_to_label.get(corpcode.norm(child_name))
            rows.append(dict(
                parent_label=parent.get("label", ""), parent_corp_code=cc,
                parent_entity_type=parent.get("entity_type", ""),
                child_name=child_name,
                child_corp_code=hit["corp_code"] if hit else "",
                child_label=hit["label"] if hit else "",
                child_match_method="target_name_exact" if hit else "unmatched",
                bsns_year=p.get("bsns_year", ""), reprt_code=p.get("reprt_code", ""),
                first_acqs_de=r.get("frst_acqs_de", ""),
                first_acqs_amount=r.get("frst_acqs_amount", ""),
                invstmnt_purps=r.get("invstmnt_purps", ""),
                bsis_qota_rt=r.get("bsis_blce_qota_rt", ""),
                trmend_qota_rt=r.get("trmend_blce_qota_rt", ""),
                trmend_qy=r.get("trmend_blce_qy", ""),
                trmend_acntbk_amount=r.get("trmend_blce_acntbk_amount", ""),
                stlm_dt=r.get("stlm_dt", ""), value_source="api",
                **prov(out_dir, m, str(r.get("rcept_no") or ""),
                       "api_row" if r.get("rcept_no") else "none")))
    rows.sort(key=lambda r: (r["parent_label"], r["bsns_year"], r["child_name"]))
    return write_csv(os.path.join(out_dir, "12_지분관계.csv"), rows,
                     ["parent_label", "parent_corp_code", "child_label", "child_name",
                      "child_corp_code", "child_match_method", "bsns_year", "reprt_code",
                      "trmend_qota_rt", "bsis_qota_rt", "first_acqs_de", "first_acqs_amount"])


def emit_documents(out_dir, metas, by_code, max_doc_bytes, doc_index=None):
    """원문 ZIP → 섹션·표 셀. 파싱은 emit 에 있으므로 파서를 고쳐도 쿼터를 다시 쓰지 않는다."""
    rows, notes = [], []
    text_dir = os.path.join(out_dir, "text")
    for m in metas:
        rcept = (m.get("params") or {}).get("rcept_no", "")
        meta_doc = (doc_index or {}).get(rcept, {})
        label = meta_doc.get("corp_label", "")
        try:
            info = docparse.read_zip(load_body(m), rcept, max_doc_bytes)
        except Exception as e:
            notes.append("%s: ZIP 열기 실패 (%s)" % (rcept, type(e).__name__))
            continue
        if info.get("truncated") or not info.get("text"):
            notes.append("%s: 본문 %s (%s) — ZIP 은 보관됨"
                         % (rcept, "크기 초과로 파싱 생략" if info.get("truncated") else "없음",
                            info.get("member_selected") or "멤버 없음"))
            continue
        if info.get("decode_replacements"):
            notes.append("%s: 인코딩 치환 %d자 (declared=%s used=%s)"
                         % (rcept, info["decode_replacements"], info["encoding_declared"],
                            info["encoding_used"]))
        sections, perr = docparse.parse_document(info["text"])
        if perr:
            notes.append("%s: %s" % (rcept, perr))

        d = os.path.join(text_dir, rcept)
        os.makedirs(d, exist_ok=True)
        with open(os.path.join(d, "_full.txt"), "w", encoding="utf-8") as f:
            f.write(info["text"])   # 섹션 매칭 실패가 내용 손실이 되지 않도록 전문을 남긴다

        base = dict(corp_label=label, corp_code=meta_doc.get("corp_code", ""),
                    report_nm=meta_doc.get("report_nm", ""), rcept_dt=meta_doc.get("rcept_dt", ""),
                    bsns_year_hint=meta_doc.get("bsns_year_hint", ""),
                    member_selected=info.get("member_selected", ""),
                    encoding_declared=info.get("encoding_declared", ""),
                    encoding_used=info.get("encoding_used", ""),
                    decode_replacements=info.get("decode_replacements", 0),
                    value_source="api")
        for si, sec, hits in docparse.match_sections(sections, config.SECTION_KEYWORDS):
            sl = docparse.slug(sec["title"])
            tp = os.path.join(d, "%03d_%s.txt" % (si, sl))
            body_text = "\n".join(sec["text_parts"])
            with open(tp, "w", encoding="utf-8") as f:
                f.write(sec["title_raw"] + "\n\n" + body_text)
            common = dict(base, section_index=si, section_title=sec["title"],
                          section_title_raw=sec["title_raw"],
                          matched_keyword="|".join(hits),
                          text_path=rel(out_dir, tp))
            rows.append(dict(common, kind="text", table_index="", row_index="",
                             cell_ord="", cell_tag="", rowspan="", colspan="",
                             unit_hint="", cell_text="",
                             text_chars=len(body_text),
                             context=body_text[:200],
                             **prov(out_dir, m, rcept, "api_row")))
            for ti, tbl in enumerate(sec["tables"]):
                xp = os.path.join(d, "table_%03d_%02d.xml" % (si, ti))
                with open(xp, "w", encoding="utf-8") as f:
                    f.write(tbl.get("raw_xml", ""))
                for ri, r in enumerate(tbl["rows"]):
                    for ci, cell in enumerate(r):
                        rows.append(dict(common, kind="table", table_index=ti,
                                         row_index=ri, cell_ord=ci, cell_tag=cell["tag"],
                                         rowspan=cell["rowspan"], colspan=cell["colspan"],
                                         unit_hint=tbl.get("unit_hint", ""),
                                         cell_text=cell["text"], text_chars="",
                                         context="", table_xml_path=rel(out_dir, xp),
                                         **prov(out_dir, m, rcept, "api_row")))
    rows.sort(key=lambda r: (r.get("corp_label") or "", r.get("rcept_no") or "",
                             r.get("section_index") or 0))
    path, _ = write_csv(os.path.join(out_dir, "11_원문추출.csv"), rows,
                        ["corp_label", "corp_code", "rcept_no", "report_nm", "rcept_dt",
                         "section_index", "section_title", "matched_keyword", "kind",
                         "table_index", "row_index", "cell_ord", "cell_tag",
                         "rowspan", "colspan", "unit_hint", "cell_text"])
    return path, notes


# ── 미확보 ────────────────────────────────────────────────────────────────
def emit_missing(out_dir, entries, years, half_years, all_metas):
    """공백의 사유를 분리해 기록한다. 뭉개면 그 자체가 삭제다."""
    got = {}
    for m in all_metas:
        p = m.get("params") or {}
        got[(m.get("endpoint"), p.get("corp_code", ""), p.get("bsns_year", ""),
             p.get("reprt_code", ""), p.get("fs_div", ""))] = m
    rows = []
    for (label, ep, cc, year, reprt, fs), skip in grid.expected_cells(entries, years, half_years).items():
        m = got.get((ep, cc, year, reprt, fs))
        if m is None:
            if skip:
                code, msg = "not_applicable_entity_window", skip.split(":", 1)[-1]
            else:
                code, msg = "not_attempted", "이 실행에서 호출되지 않음(범위·한도·중단)"
            rows.append(dict(corp_label=label, corp_code=cc, source_endpoint=ep,
                             bsns_year=year, reprt_code=reprt, fs_div=fs,
                             reason_code=code, api_status="", api_message=msg,
                             retryable="N" if code == "not_applicable_entity_window" else "Y",
                             raw_path="", fetched_at=""))
            continue
        st = m.get("api_status", "")
        n = len((load_json(m) or {}).get("list") or []) if config.ENDPOINTS.get(ep, {}).get("kind") == "json" else 1
        if st == "000" and n > 0:
            continue
        if st == "000" and n == 0:
            code = "api_000_empty_list"
        elif st == "013":
            code = "api_013"
        elif st in config.FATAL_STATUS:
            code = "fatal_status"
        elif st == "TRANSPORT":
            code = "transport_failed"
        else:
            code = "schema_mismatch" if st else "parse_failed"
        rows.append(dict(corp_label=label, corp_code=cc, source_endpoint=ep,
                         bsns_year=year, reprt_code=reprt, fs_div=fs, reason_code=code,
                         api_status=st, api_message=m.get("api_message", ""),
                         retryable="Y" if code in ("fatal_status", "transport_failed") else "N",
                         raw_path=rel(out_dir, m["_body_path"]),
                         fetched_at=m.get("fetched_at", "")))
    rows.sort(key=lambda r: (r["corp_label"], r["source_endpoint"], r["bsns_year"]))
    return write_csv(os.path.join(out_dir, "99_미확보목록.csv"), rows,
                     ["corp_label", "corp_code", "source_endpoint", "bsns_year",
                      "reprt_code", "fs_div", "reason_code", "api_status", "api_message",
                      "retryable", "fetched_at", "raw_path"])


# ── XLSX ──────────────────────────────────────────────────────────────────
XLSX_MAX_ROWS = 1000000
XLSX_MAX_CELL = 32000


def emit_xlsx(out_dir, csv_paths):
    try:
        from openpyxl import Workbook
    except ImportError:
        print("  [건너뜀] openpyxl 이 없어 XLSX 를 만들지 않습니다 "
              "(pip install -r scripts/dart/requirements.txt). CSV 가 정본입니다.")
        return None
    wb = Workbook(write_only=True)
    notes = []
    for p in csv_paths:
        if not p or not os.path.exists(p):
            continue
        name = os.path.splitext(os.path.basename(p))[0][:31]
        with open(p, encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f)
            header = next(reader, None)
            if header is None:
                continue
            ws = wb.create_sheet(name)
            ws.append(header)
            n, part = 1, 1
            for row in reader:
                if n >= XLSX_MAX_ROWS:
                    part += 1
                    notes.append("%s → %s_p%d 로 분할" % (name, name[:27], part))
                    ws = wb.create_sheet(("%s_p%d" % (name[:27], part))[:31])
                    ws.append(header)
                    n = 1
                # 전 셀 문자열: corp_code 선행 0 과 rcept_no 14자리가 숫자로 뭉개지지 않게
                ws.append([(c[:XLSX_MAX_CELL] if len(c) > XLSX_MAX_CELL else c) for c in row])
                n += 1
    if notes:
        ws = wb.create_sheet("_분할안내")
        ws.append(["원본 시트", "안내"])
        for t in notes:
            ws.append([t, "행 수 초과로 분할됨 — 잘라내지 않았습니다"])
    path = os.path.join(out_dir, "DART_추출결과.xlsx")
    wb.save(path)
    return path


# ── 실행 보고서 ───────────────────────────────────────────────────────────
def write_run_report(out_dir, info):
    p = os.path.join(out_dir, "RUN_REPORT.md")
    L = ["# 수집 실행 보고서", "", "생성: %s" % datetime.now().astimezone().isoformat(timespec="seconds"), ""]
    for title, items in info:
        L.append("## %s" % title)
        if not items:
            L.append("- 없음")
        else:
            L += ["- %s" % x for x in items]
        L.append("")
    with open(p, "w", encoding="utf-8") as f:
        f.write("\n".join(L))
    return p


# ── 엔트리포인트 ──────────────────────────────────────────────────────────
def emit_all(out_dir, years=None, half_years=None, max_doc_bytes=docparse.DEFAULT_MAX_DOC_BYTES):
    years = years or config.DEFAULT_YEARS
    half_years = half_years or config.DEFAULT_HALF_YEARS
    by_code, entries = corp_lookup(out_dir)
    all_metas = scan_raw(out_dir)
    by_ep = {}
    for m in all_metas:
        by_ep.setdefault(m["_endpoint"], []).append(m)

    produced, report = [], []
    p, _ = emit_company(out_dir, by_ep.get("company", []), by_code)
    produced.append(p)

    lp, list_rows, pag_bad = emit_list(out_dir, by_ep.get("list", []), by_code)
    produced.append(lp)

    fin_paths, drift = emit_financials(out_dir, by_ep.get("fnlttSinglAcntAll", []), by_code)
    produced = [fin_paths[2]] + produced + fin_paths[:2]

    for ep in config.REPORT_ENDPOINTS:
        produced.append(emit_generic(out_dir, ep, by_ep.get(ep, []), by_code,
                                     config.ENDPOINTS[ep]["csv"])[0])

    doc_index = {}
    for r in list_rows:
        doc_index.setdefault(r.get("rcept_no", ""), dict(
            corp_label=r.get("corp_label", ""), corp_code=r.get("corp_code", ""),
            report_nm=r.get("report_nm", ""), rcept_dt=r.get("rcept_dt", "")))
    dp, dnotes = emit_documents(out_dir, by_ep.get("document", []), by_code,
                                max_doc_bytes, doc_index)
    produced.append(dp)
    produced.append(emit_ownership(out_dir, by_ep.get("otrCprInvstmntSttus", []),
                                   by_code, entries)[0])
    produced.append(emit_missing(out_dir, entries, years, half_years, all_metas)[0])

    report.append(("스키마 드리프트 (예상 밖 응답 키)",
                   ["`%s` — 컬럼으로는 보존됨. 의미 확인 필요" % d for d in drift]))
    report.append(("페이징 불일치 (조회 중 total_count 변동)",
                   ["corp_code=%s %s~%s: total_count %s" % (k[0], k[1], k[2], sorted(v))
                    for k, v in pag_bad.items()]))
    report.append(("원문 파싱 경고", dnotes))
    report.append(("정체성 검증 실패",
                   ["%s: %s" % (e["label"], e.get("identity_verified_by", ""))
                    for e in entries if e.get("identity_check") == "불일치"]))
    rp = write_run_report(out_dir, report)

    produced = [x for x in produced if x]
    xp = emit_xlsx(out_dir, produced)
    return produced + [rp] + ([xp] if xp else [])
